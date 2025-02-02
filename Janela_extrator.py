import pyodbc
import pandas as pd

import os

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

from tkinter import *
from tkinter import messagebox
from tkcalendar import DateEntry
import tkinter as tk
from tkinter import ttk


#cria a conecxão ao banco
def get_db_connection():
    try:
        conn = pyodbc.connect(
            'DRIVER={SQL Server};'
            'SERVER=servidor;'
            'DATABASE=nome do banco;'
            'UID=login banco;'
            'PWD=senha banco;'
            'Trusted_Connection=no;'
        )
        return conn
    #checa a conexão e retorna em caso de erro
    except pyodbc.Error as e:
        print("Erro na conexão:", e)
        return None

#conecta ao banco
def consultar_bd(fornecedor, localizacao, periodo_In, periodo_fin):
    conn = get_db_connection()
    #checa e retorna caso a conexão falhe
    if conn is None:
        print("Falha na conexão a banco de dados.")
        return []
    #executa o SQL
    try:
        cur = conn.cursor()
        query = '''
        SELECT 
        LOCALIZACAO, 
        DATA,
        CLI_NOME,
        CLI_TPPESSOA,
        CLI_CPF,
        UF,
        CIDADE,
        NATUREZA,
        VENDEDOR,
        PRODUTO,
        QUANTIDADE, 
        VALOR_TOTAL
    FROM VW_FATURAMENTO_DETALHADO
    JOIN CLIENTE C ON CLI_CODI = COD_CLIENTE_GERAL
    WHERE FOR_CODI = ?
    AND NATUREZA IN ('VEN', 'BOV', 'DVE')
    AND LOCALIZACAO = ?
    AND CONVERT(DATETIME, DATA, 103) BETWEEN ? AND ?
    ORDER BY DATA DESC
        '''
         # Formatando corretamente as listas de parâmetros
        query = query.format() 
        parametros = [fornecedor] + [localizacao] + [periodo_In, periodo_fin]
        cur.execute(query, parametros)

        #retorna os dados recolhidos no select
        rows1 = cur.fetchall()
        cur.close()
        conn.close()
        return rows1
    #retorna em caso de erro no recolhimento dos dados
    except pyodbc.Error as e:
        print("Erro ao juntar os dados data:", e)
        return []

def consultar_bd_estoque(fornecedor):
    conn = get_db_connection()
    if conn is None:
        print("Falha na conexão ao banco de dados.")
        return []
    
    try:
        cur = conn.cursor()
        query = '''
        SELECT P.PRO_CODI AS CODIGO,
               SEC_CODI AS SECAO,
               PRO_DESC AS PRODUTO,
	           CAST(EST_QUAN AS INT) AS ESTOQUE
        FROM [Gestor].[dbo].[PRODUTO] P
        JOIN dbo.ESTOQUE M ON M.PRO_CODI = P.PRO_CODI
        WHERE FOR_CODI = ?
        AND SEC_CODI IN ('009','012')
        AND EST_QUAN != '0'
        ORDER BY PRO_DESC ASC
        '''
        cur.execute(query, [fornecedor])
        rows2 = cur.fetchall()
        cur.close()
        conn.close()
        return rows2
    
    except pyodbc.Error as e:
        print("Erro ao consultar o estoque:", e)
        return []

#botão consultar recolhendo os dados da caixa de entrada   
def on_consulta():
    fornecedor = forn_var.get()
    natureza = nat_var.get().split(',')  # Separando a string em lista
    localizacao = loc_var.get()
    periodo_In = periodo_ini_var.get()
    periodo_fin = periodo_fin_var.get()
    resultados = consultar_bd(fornecedor, natureza, localizacao, periodo_In, periodo_fin)
    exibir_resultados(resultados)  # Passa os resultados para exibir na função exibir_resultados
      
#botão salvar em excel recolhendo os dados da caixa de entrada   
def on_salva():
    fornecedor = forn_var.get()
    natureza = nat_var.get().split(',')  # Separando a string em lista
    localizacao = loc_var.get()
    periodo_In = periodo_ini_var.get()
    periodo_fin = periodo_fin_var.get()
    resultados = consultar_bd(fornecedor, natureza, localizacao, periodo_In, periodo_fin)
    filename = filename_var.get()
    if not filename:
        messagebox.showwarning("Atenção", "Por favor, insira o nome para o arquivo")
        return
    filepath = f"extraido/{filename}.xlsx"
    save_to_excel(resultados, filepath) # Passando resultados para serem salvos na função save_to_excel

#exibe os 5 primeiros resultados para validação prévia
def exibir_resultados(resultados):
    janela_resultados = Toplevel(root)
    janela_resultados.title("Resultados da Consulta")
     
    for i in range(min(5, len(resultados))):
        row = resultados[i]
        ttk.Label(janela_resultados, text=f"Resultado {i+1}: {row}").grid(column=0, row=i, padx=10, pady=5)
    
    global resultados_df
    colunas = [
        'DISTRIBUIDOR', 'DATA FATURAMENTO', 'RAZÃO SOCIAL', 'CNPJ', 'ESTADO',
        'MUNICÍPIO', 'NATUREZA', 'VENDEDOR', 'DESCR ITEM', 'QTDE', 'VALOR TOTAL'
    ]
    resultados_df = pd.DataFrame(resultados, columns=colunas)

# Salva os dados em um arquivo Excel com duas abas, se necessário
def save_to_excel(rows1, rows2, filename):
    os.makedirs(os.path.dirname(filename), exist_ok=True)
    
    data1 = [tuple(row) for row in rows1]
    df1 = pd.DataFrame(data1, columns=[
        'DISTRIBUIDOR', 'DATA FATURAMENTO', 'RAZÃO SOCIAL', 'CNPJ', 'ESTADO', 'MUNICÍPIO', 'NATUREZA', 'VENDEDOR', 'DESCR ITEM', 'QTDE', 'VALOR TOTAL BRUTO'])
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df1.to_excel(writer, sheet_name='Faturamento', index=False)
        
        if rows2:
            data2 = [tuple(row) for row in rows2]
            df2 = pd.DataFrame(data2, columns=[
                'CODIGO', 'SECAO', 'PRODUTO', 'ESTOQUE'])
            df2.to_excel(writer, sheet_name='Estoque', index=False)
    
    workbook = load_workbook(filename)
    sheet1 = workbook['Faturamento']
    red_fill = PatternFill(start_color="FFFF5050", end_color="FFFF5050", fill_type="solid")
    white_font = Font(color="FFFFFF", name="Aptos Narrow", bold=True)
    for cell in sheet1[1]:
        cell.fill = red_fill
        cell.font = white_font
    
    sheet1.auto_filter.ref = sheet1.dimensions
    
    if rows2:
        sheet2 = workbook['Estoque']
        red_fill = PatternFill(start_color="FFFF5050", end_color="FFFF5050", fill_type="solid")
        for cell in sheet2[1]:
            cell.fill = red_fill
            cell.font = white_font
        sheet2.auto_filter.ref = sheet2.dimensions

    workbook.save(filename)
    print(f"Aqruivo salvo como: {filename}")
    messagebox.showinfo("Successo", f"Aqruivo salvo como: {filename}")
 
# Botão enviar email
def on_enviar():
    smtp_server = 'smtps.uhserver.com'
    smtp_port = 465
    smtp_user = 'ti@veteagro.com.br'
    smtp_password = 'Veteagro@16'
    from_addr = 'ti@veteagro.com.br'
    to_addr = email_var.get()  # Obtém o e-mail digitado na caixa de texto

    #retorna caso a caixa de email esteja vazia
    if not to_addr:
        messagebox.showwarning("Atenção", "Por favor, insira o e-mail do destinatário.")
        return

    subject = f'RELATORIO SELL OUT {filename_var.get()}'
    body = f'Segue em anexo relatório {filename_var.get()} para análise e redirecionamento'
    file_path = f'extraido/{filename_var.get()}.xlsx'
    
    send_email_with_attachment(smtp_server, smtp_port, smtp_user, smtp_password, from_addr, to_addr, subject, body, file_path)
# Envia o email
def send_email_with_attachment(smtp_server, smtp_port, smtp_user, smtp_password, from_addr, to_addr, subject, body, file_path):
    try:
        # Cria a mensagem
        msg = MIMEMultipart()
        msg['From'] = from_addr
        msg['To'] = to_addr
        msg['Subject'] = subject

        # Adiciona o corpo do email
        msg.attach(MIMEText(body, 'plain'))

        # Anexa o arquivo
        with open(file_path, 'rb') as attachment:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename={os.path.basename(file_path)}')
            msg.attach(part)

        # Configura o servidor SMTP e envia o email
        with smtplib.SMTP_SSL(smtp_server, smtp_port) as server:
            server.login(smtp_user, smtp_password)
            server.sendmail(from_addr, to_addr, msg.as_string())

        print("Email enviado com sucesso!")
        messagebox.showinfo("Sucesso", "Email enviado com sucesso!")

    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao enviar o email: {e}")

# Inicia a interface gráfica
root = tk.Tk()
root.title("Consulta ao Banco de Dados")

# Listas predefinidas
fornecedores = {
    'ZOETIS': '00155',
    'CEVA': '00116',
    'ELANCO': '00290',
    'KONIG': '01076',
    'VETOQUINOL': '00993',
    'AGENER': '00790',
    'DECHRA': '01069',
    'BST-VET&AGRO': '00207',
    'GEA': '00224',
    'LABYES': '00262',
    'OURO FINO': '00940'
}

loc = {'MATRIZ':'001', 'LOJA' :'002', 'INDUSTRIA':'004'}

# Variáveis para armazenar as seleções
forn_var = tk.StringVar()
loc_var = tk.StringVar()
periodo_ini_var = tk.StringVar()
periodo_fin_var = tk.StringVar()
filename_var = StringVar()
email_var = StringVar()
gerar_estoque_var = tk.BooleanVar()

def on_fornecedor_select(event):
    nome_selecionado = fornecedor_combobox.get()
    numero_fornecedor = fornecedores[nome_selecionado]
    forn_var.set(numero_fornecedor)

# Preencher o combobox com os nomes dos fornecedores
Label(root, text="Fornecedor:").grid(column=0, row=0, padx=10, pady=5)
fornecedor_combobox = ttk.Combobox(root, values=list(fornecedores.keys()))
fornecedor_combobox.grid(column=1, row=0, padx=10, pady=5)
fornecedor_combobox.bind("<<ComboboxSelected>>", on_fornecedor_select)

def on_loc_select(event):
    num_selecionado = local_combobox.get()
    numero_loc = loc[num_selecionado]
    loc_var.set(numero_loc)

# Preencher o combobox com os nomes da localizacao
Label(root, text="Localização:").grid(column=0, row=1, padx=10, pady=5)
local_combobox = ttk.Combobox(root, values=list(loc.keys()))
local_combobox.grid(column=1, row=1, padx=10, pady=5)
local_combobox.bind("<<ComboboxSelected>>", on_loc_select)

#periodo inicial
ttk.Label(root, text="Data Inicial:").grid(column=0, row=4, padx=10, pady=5)
data_ini_entry = DateEntry(root, textvariable=periodo_ini_var, date_pattern='yyyy-mm-dd')
data_ini_entry.grid(column=1, row=4, padx=10, pady=5)

#periodo final
ttk.Label(root, text="Data Final:").grid(column=0, row=5, padx=10, pady=5)
data_fin_entry = DateEntry(root, textvariable=periodo_fin_var, date_pattern='yyyy-mm-dd')
data_fin_entry.grid(column=1, row=5, padx=10, pady=5)

# Botão para realizar a consulta
ttk.Button(root, text="Consultar", command=on_consulta).grid(column=1, row=6, padx=10, pady=5)


# Label e Entry para o nome do arquivo
Label(root, text="Nome do arquivo:").grid(column=0, row=7, padx=10, pady=5)
# Variável para armazenar o estado da caixa de seleção
Entry(root, textvariable=filename_var).grid(column=1, row=7, padx=10, pady=5)

# Adicione a Checkbutton para gerar o estoque
ttk.Checkbutton(root, text="Gerar aba de Estoque", variable=gerar_estoque_var).grid(column=0, row=8, padx=10, pady=5)

# Botão para salvar
ttk.Button(root, text="Salvar", command=on_salva).grid(column=1, row=8, padx=10, pady=5)

# Entry para o email a ser enviado com o arquivo
ttk.Label(root, text="deseja enviar o arquivo por email?").grid(column=0, row=9, padx=10, pady=5)
ttk.Entry(root, textvariable=email_var).grid(column=1, row=9, padx=10, pady=5)
# Botão para enviar
ttk.Button(root, text="enviar", command=on_enviar).grid(column=0, row=10, padx=10, pady=5)
ttk.Button(root, text="cancelar", command=root.destroy).grid(column=1, row=10, padx=10, pady=5)

# Executa a interface gráfica
root.mainloop()
